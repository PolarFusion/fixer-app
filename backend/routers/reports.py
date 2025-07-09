"""
Роутер для генерации отчетов в форматах PDF и XLSX.
Поддерживает индивидуальные отчеты по заявкам и сводные отчеты.
"""

import os
import io
from datetime import datetime, timedelta
from typing import Annotated, List
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

# Импорты для генерации отчетов
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from database import get_session, get_ticket_by_id, get_tickets_for_user
from models import Ticket, TicketStatus, User, UserRole, DigestRange
from routers.auth import get_current_active_user, check_admin_role

# ===== КОНФИГУРАЦИЯ =====

router = APIRouter(prefix="/api/reports", tags=["reports"])

# Настройки для PDF
try:
    # Попытка загрузить русский шрифт (если доступен)
    # В реальном проекте добавьте .ttf файл русского шрифта
    pass
except:
    pass


# ===== ГЕНЕРАЦИЯ PDF =====

def create_ticket_pdf(ticket: Ticket) -> io.BytesIO:
    """Создает PDF отчет для одной заявки"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Создаем стили для русского текста
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Центрирование
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=12
    )
    
    story = []
    
    # Заголовок
    story.append(Paragraph(f"ОТЧЕТ ПО ЗАЯВКЕ №{ticket.id}", title_style))
    story.append(Spacer(1, 20))
    
    # Основная информация
    info_data = [
        ['Заголовок:', ticket.title],
        ['Адрес:', ticket.address],
        ['Описание:', ticket.description],
        ['Статус:', ticket.status.value],
        ['Приоритет:', f"{ticket.priority}/5"],
        ['Дата создания:', ticket.created_at.strftime('%d.%m.%Y %H:%M')],
        ['Крайний срок:', ticket.deadline.strftime('%d.%m.%Y %H:%M')],
    ]
    
    # Добавляем информацию о пользователях
    if ticket.customer:
        info_data.append(['Заказчик:', ticket.customer.full_name])
    if ticket.executor:
        info_data.append(['Исполнитель:', ticket.executor.full_name])
    
    # Временная информация
    if ticket.started_at:
        info_data.append(['Начата:', ticket.started_at.strftime('%d.%m.%Y %H:%M')])
    if ticket.completed_at:
        info_data.append(['Завершена:', ticket.completed_at.strftime('%d.%m.%Y %H:%M')])
        # Вычисляем время выполнения
        if ticket.started_at:
            duration = ticket.completed_at - ticket.started_at
            hours = duration.total_seconds() / 3600
            info_data.append(['Время выполнения:', f"{hours:.1f} часов"])
    
    # Комментарии
    if ticket.completion_comment:
        info_data.append(['Комментарий:', ticket.completion_comment])
    if ticket.rejection_reason:
        info_data.append(['Причина отклонения:', ticket.rejection_reason])
    
    # Создаем таблицу
    table = Table(info_data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Информация о фотографиях
    if ticket.before_photo_path or ticket.after_photo_path:
        story.append(Spacer(1, 20))
        story.append(Paragraph("ПРИЛОЖЕННЫЕ ФОТОГРАФИИ:", styles['Heading2']))
        
        photos_info = []
        if ticket.before_photo_path:
            photos_info.append(f"• Фото 'до': {os.path.basename(ticket.before_photo_path)}")
        if ticket.after_photo_path:
            photos_info.append(f"• Фото 'после': {os.path.basename(ticket.after_photo_path)}")
        
        for photo_info in photos_info:
            story.append(Paragraph(photo_info, normal_style))
    
    # Подвал
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        styles['Normal']
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def create_digest_pdf(tickets: List[Ticket], period: str) -> io.BytesIO:
    """Создает сводный PDF отчет"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Заголовок
    period_text = "дневной" if period == "daily" else "недельный"
    title = f"СВОДНЫЙ {period_text.upper()} ОТЧЕТ"
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 20))
    
    # Общая статистика
    total_tickets = len(tickets)
    status_counts = {}
    for status in TicketStatus:
        status_counts[status] = len([t for t in tickets if t.status == status])
    
    stats_data = [
        ['Показатель', 'Значение'],
        ['Всего заявок:', str(total_tickets)],
        ['Ожидающих:', str(status_counts.get(TicketStatus.PENDING, 0))],
        ['В работе:', str(status_counts.get(TicketStatus.IN_PROGRESS, 0))],
        ['Выполненных:', str(status_counts.get(TicketStatus.DONE, 0))],
        ['Отклоненных:', str(status_counts.get(TicketStatus.REJECTED, 0))],
    ]
    
    # Средняя скорость выполнения
    completed_tickets = [t for t in tickets if t.status == TicketStatus.DONE and t.started_at and t.completed_at]
    if completed_tickets:
        avg_time = sum(
            (t.completed_at - t.started_at).total_seconds() 
            for t in completed_tickets
        ) / len(completed_tickets) / 3600
        stats_data.append(['Среднее время выполнения:', f"{avg_time:.1f} часов"])
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(Paragraph("ОБЩАЯ СТАТИСТИКА", styles['Heading2']))
    story.append(stats_table)
    story.append(Spacer(1, 20))
    
    # Список заявок
    if tickets:
        story.append(Paragraph("СПИСОК ЗАЯВОК", styles['Heading2']))
        
        tickets_data = [['№', 'Заголовок', 'Статус', 'Исполнитель', 'Дата создания']]
        
        for ticket in tickets[:20]:  # Ограничиваем 20 заявками
            executor_name = ticket.executor.full_name if ticket.executor else "Не назначен"
            tickets_data.append([
                str(ticket.id),
                ticket.title[:30] + "..." if len(ticket.title) > 30 else ticket.title,
                ticket.status.value,
                executor_name[:20] + "..." if len(executor_name) > 20 else executor_name,
                ticket.created_at.strftime('%d.%m.%Y')
            ])
        
        tickets_table = Table(tickets_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.5*inch, 1*inch])
        tickets_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(tickets_table)
    
    # Подвал
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        styles['Normal']
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


# ===== ГЕНЕРАЦИЯ XLSX =====

def create_ticket_xlsx(ticket: Ticket) -> io.BytesIO:
    """Создает XLSX отчет для одной заявки"""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заявка №{ticket.id}"
    
    # Стили
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Заголовок
    ws.merge_cells('A1:B1')
    ws['A1'] = f"ОТЧЕТ ПО ЗАЯВКЕ №{ticket.id}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Данные
    row = 3
    data_pairs = [
        ('Заголовок', ticket.title),
        ('Адрес', ticket.address),
        ('Описание', ticket.description),
        ('Статус', ticket.status.value),
        ('Приоритет', f"{ticket.priority}/5"),
        ('Дата создания', ticket.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Крайний срок', ticket.deadline.strftime('%d.%m.%Y %H:%M')),
    ]
    
    if ticket.customer:
        data_pairs.append(('Заказчик', ticket.customer.full_name))
    if ticket.executor:
        data_pairs.append(('Исполнитель', ticket.executor.full_name))
    if ticket.started_at:
        data_pairs.append(('Начата', ticket.started_at.strftime('%d.%m.%Y %H:%M')))
    if ticket.completed_at:
        data_pairs.append(('Завершена', ticket.completed_at.strftime('%d.%m.%Y %H:%M')))
        if ticket.started_at:
            duration = ticket.completed_at - ticket.started_at
            hours = duration.total_seconds() / 3600
            data_pairs.append(('Время выполнения', f"{hours:.1f} часов"))
    
    if ticket.completion_comment:
        data_pairs.append(('Комментарий', ticket.completion_comment))
    if ticket.rejection_reason:
        data_pairs.append(('Причина отклонения', ticket.rejection_reason))
    
    for label, value in data_pairs:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = value
        row += 1
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_digest_xlsx(tickets: List[Ticket], period: str) -> io.BytesIO:
    """Создает сводный XLSX отчет"""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Лист со статистикой
    ws_stats = wb.active
    ws_stats.title = "Статистика"
    
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    
    ws_stats['A1'] = f"СВОДНЫЙ {'ДНЕВНОЙ' if period == 'daily' else 'НЕДЕЛЬНЫЙ'} ОТЧЕТ"
    ws_stats['A1'].font = header_font
    ws_stats.merge_cells('A1:B1')
    
    # Статистика
    total_tickets = len(tickets)
    status_counts = {}
    for status in TicketStatus:
        status_counts[status] = len([t for t in tickets if t.status == status])
    
    stats_data = [
        ('Всего заявок', total_tickets),
        ('Ожидающих', status_counts.get(TicketStatus.PENDING, 0)),
        ('В работе', status_counts.get(TicketStatus.IN_PROGRESS, 0)),
        ('Выполненных', status_counts.get(TicketStatus.DONE, 0)),
        ('Отклоненных', status_counts.get(TicketStatus.REJECTED, 0)),
    ]
    
    row = 3
    for label, value in stats_data:
        ws_stats[f'A{row}'] = label
        ws_stats[f'A{row}'].font = label_font
        ws_stats[f'B{row}'] = value
        row += 1
    
    # Лист с заявками
    ws_tickets = wb.create_sheet("Заявки")
    headers = ['ID', 'Заголовок', 'Адрес', 'Статус', 'Приоритет', 'Заказчик', 'Исполнитель', 'Дата создания', 'Завершена']
    
    for col, header in enumerate(headers, 1):
        cell = ws_tickets.cell(row=1, column=col)
        cell.value = header
        cell.font = label_font
    
    for row, ticket in enumerate(tickets, 2):
        ws_tickets.cell(row=row, column=1, value=ticket.id)
        ws_tickets.cell(row=row, column=2, value=ticket.title)
        ws_tickets.cell(row=row, column=3, value=ticket.address)
        ws_tickets.cell(row=row, column=4, value=ticket.status.value)
        ws_tickets.cell(row=row, column=5, value=ticket.priority)
        ws_tickets.cell(row=row, column=6, value=ticket.customer.full_name if ticket.customer else "")
        ws_tickets.cell(row=row, column=7, value=ticket.executor.full_name if ticket.executor else "")
        ws_tickets.cell(row=row, column=8, value=ticket.created_at.strftime('%d.%m.%Y %H:%M'))
        ws_tickets.cell(row=row, column=9, value=ticket.completed_at.strftime('%d.%m.%Y %H:%M') if ticket.completed_at else "")
    
    # Автоширина
    for column in ws_tickets.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_tickets.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ===== ЭНДПОИНТЫ =====

@router.get("/{ticket_id}")
async def get_ticket_report(
    ticket_id: int,
    format: str = "pdf",  # pdf или xlsx
    current_user: Annotated[User, Depends(get_current_active_user)] = Depends(),
    session: AsyncSession = Depends(get_session)
):
    """Генерация отчета по конкретной заявке"""
    # Получаем заявку
    ticket = await get_ticket_by_id(session, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    
    # Проверяем права доступа
    if current_user.role == UserRole.CUSTOMER and ticket.customer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    elif current_user.role == UserRole.EXECUTOR and ticket.executor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    
    # Генерируем отчет
    if format.lower() == "pdf":
        buffer = create_ticket_pdf(ticket)
        media_type = "application/pdf"
        filename = f"ticket_{ticket_id}_report.pdf"
    elif format.lower() == "xlsx":
        buffer = create_ticket_xlsx(ticket)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"ticket_{ticket_id}_report.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Неподдерживаемый формат")
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/digest/{range}")
async def get_digest_report(
    range: DigestRange,
    format: str = "pdf",
    current_user: Annotated[User, Depends(check_admin_role)] = Depends(),
    session: AsyncSession = Depends(get_session)
):
    """Генерация сводного отчета (только для админов)"""
    # Определяем период
    now = datetime.utcnow()
    if range == DigestRange.DAILY:
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    else:  # weekly
        start_date = now - timedelta(days=7)
    
    # Получаем заявки за период
    query = select(Ticket).where(Ticket.created_at >= start_date)
    result = await session.exec(query)
    tickets = result.all()
    
    # Получаем полную информацию о заявках
    full_tickets = []
    for ticket in tickets:
        full_ticket = await get_ticket_by_id(session, ticket.id)
        if full_ticket:
            full_tickets.append(full_ticket)
    
    # Генерируем отчет
    period_name = "daily" if range == DigestRange.DAILY else "weekly"
    
    if format.lower() == "pdf":
        buffer = create_digest_pdf(full_tickets, period_name)
        media_type = "application/pdf"
        filename = f"digest_{period_name}_{now.strftime('%Y%m%d')}.pdf"
    elif format.lower() == "xlsx":
        buffer = create_digest_xlsx(full_tickets, period_name)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"digest_{period_name}_{now.strftime('%Y%m%d')}.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Неподдерживаемый формат")
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )